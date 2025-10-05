import streamlit as st
import pandas as pd
import gspread # Google Sheets APIé€£æºç”¨
from oauth2client.service_account import ServiceAccountCredentials # èªè¨¼æƒ…å ±
import json # JSONã‚­ãƒ¼ãƒ•ã‚¡ã‚¤ãƒ«èª­ã¿è¾¼ã¿ç”¨
from io import BytesIO # Excelã‚¢ãƒƒãƒ—ãƒ­ãƒ¼ãƒ‰/ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ç”¨
from openpyxl import load_workbook # Excelèª­ã¿æ›¸ãç”¨

st.set_page_config(
    page_title="Googleãƒ•ã‚©ãƒ¼ãƒ ã‹ã‚‰Excelæˆæ¥­ã‚«ãƒ¼ãƒ‰ä½œæˆ",
    page_icon="ğŸ“",
    layout="centered"
)

st.title("ğŸ“ Googleãƒ•ã‚©ãƒ¼ãƒ ã‹ã‚‰Excelæˆæ¥­ã‚«ãƒ¼ãƒ‰ã‚’ä½œæˆ")
st.markdown("---")

# Google Sheets APIã‹ã‚‰ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—ã™ã‚‹é–¢æ•°
@st.cache_data(ttl=3600)
def load_data_from_google_sheet(spreadsheet_name, worksheet_name):
    try:
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        
        creds_json_string = st.secrets.get("GOOGLE_SHEETS_CREDENTIALS")
        
        if creds_json_string is None:
            st.error(
                "Streamlitã®secretsã« 'GOOGLE_SHEETS_CREDENTIALS' ã‚­ãƒ¼ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã€‚\n"
                "ã‚µãƒ¼ãƒ“ã‚¹ã‚¢ã‚«ã‚¦ãƒ³ãƒˆã®JSONã‚­ãƒ¼ã‚’ `secrets.toml` ã¾ãŸã¯ Streamlit Cloudã®è¨­å®šã§ `GOOGLE_SHEETS_CREDENTIALS` "
                "ã¨ã„ã†åå‰ã§è¨­å®šã—ã¦ãã ã•ã„ã€‚"
            )
            st.stop()
        
        creds_info = json.loads(creds_json_string)
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_info, scope)
        client = gspread.authorize(creds)
        
        spreadsheet = client.open(spreadsheet_name)
        worksheet = spreadsheet.worksheet(worksheet_name)
        
        data = worksheet.get_all_values()
        
        # ãƒ˜ãƒƒãƒ€ãƒ¼è¡Œã‚’ç‰¹å®šã—ã€ãã®å¾Œã®ãƒ‡ãƒ¼ã‚¿ã‚’è¾æ›¸ã®ãƒªã‚¹ãƒˆã¨ã—ã¦å‡¦ç†
        if not data:
            return []

        headers = data[0]
        records = data[1:]
        
        processed_records = []
        for row in records:
            if any(cell.strip() for cell in row): # ç©ºè¡Œã‚’ã‚¹ã‚­ãƒƒãƒ—
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row): # åˆ—ã®ã‚¤ãƒ³ãƒ‡ãƒƒã‚¯ã‚¹ãŒç¯„å›²å†…ã§ã‚ã‚‹ã“ã¨ã‚’ç¢ºèª
                        value = str(row[i]).strip()
                        # ãƒªã‚¹ãƒˆã¨ã—ã¦æ‰±ã†é …ç›® (ã‚»ãƒŸã‚³ãƒ­ãƒ³åŒºåˆ‡ã‚Š)
                        if header in ['å°å…¥ã®æµã‚Œ', 'æ´»å‹•ã®æµã‚Œ', 'æŒ¯ã‚Šè¿”ã‚Šã®æµã‚Œ', 'æŒ‡å°ã®ãƒã‚¤ãƒ³ãƒˆ', 'æ•™æå†™çœŸURL']:
                            row_dict[header] = [item.strip() for item in value.split(';') if item.strip()] if value else []
                        # ãƒãƒƒã‚·ãƒ¥ã‚¿ã‚° (ã‚«ãƒ³ãƒåŒºåˆ‡ã‚Š)
                        elif header == 'ãƒãƒƒã‚·ãƒ¥ã‚¿ã‚°':
                            row_dict[header] = [item.strip() for item in value.split(',') if item.strip()] if value else []
                        # æ•°å€¤ã«å¤‰æ›ã™ã‚‹é …ç›® (ã‚¨ãƒ©ãƒ¼ãƒãƒ³ãƒ‰ãƒªãƒ³ã‚°ä»˜ã)
                        elif header == 'å˜å…ƒå†…ã§ã®ä¸¦ã³é †':
                            try:
                                row_dict[header] = int(value)
                            except (ValueError, TypeError):
                                row_dict[header] = 9999 # ãƒ‡ãƒ•ã‚©ãƒ«ãƒˆå€¤
                        # ICTæ´»ç”¨æœ‰ç„¡ (ãƒ–ãƒ¼ãƒ«å€¤ã¾ãŸã¯æ–‡å­—åˆ—)
                        elif header == 'ICTæ´»ç”¨æœ‰ç„¡':
                            val_lower = value.lower()
                            if val_lower == 'true' or val_lower == 'ã¯ã„':
                                row_dict[header] = 'ã‚ã‚Š'
                            elif val_lower == 'false' or val_lower == 'ã„ã„ãˆ':
                                row_dict[header] = 'ãªã—'
                            else:
                                row_dict[header] = value # ãã®ä»–ã®å ´åˆã¯ãã®ã¾ã¾
                        else:
                            row_dict[header] = value
                    else:
                        row_dict[header] = '' # ãƒ‡ãƒ¼ã‚¿ãŒãªã„å ´åˆã€ç©ºæ–‡å­—åˆ—ã‚’è¨­å®š

                # ãƒ¦ãƒ‹ãƒ¼ã‚¯ãªIDã‚’ä»˜ä¸ (Google Sheetsç”±æ¥ã§ã‚ã‚‹ã“ã¨ã‚’ç¤ºã™)
                # ã‚¿ã‚¤ãƒ ã‚¹ã‚¿ãƒ³ãƒ—ã‚’IDã®ä¸€éƒ¨ã¨ã—ã¦ä½¿ã†ã“ã¨ã§ã€ã‚ˆã‚Šãƒ¦ãƒ‹ãƒ¼ã‚¯æ€§ã‚’é«˜ã‚ã‚‹
                timestamp = row_dict.get('ã‚¿ã‚¤ãƒ ã‚¹ã‚¿ãƒ³ãƒ—', f"no_timestamp_{len(processed_records)}")
                row_dict['generated_id'] = f"gs_{timestamp}_{len(processed_records)}"
                processed_records.append(row_dict)
        
        return processed_records

    except KeyError as e:
        st.error(f"Google Sheets APIã®èªè¨¼æƒ…å ±ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã€‚`secrets.toml`ã¾ãŸã¯Streamlit Cloudã®è¨­å®šã‚’ç¢ºèªã—ã¦ãã ã•ã„: {e}")
        st.stop()
    except Exception as e:
        st.error(f"Googleã‚¹ãƒ—ãƒ¬ãƒƒãƒ‰ã‚·ãƒ¼ãƒˆã‹ã‚‰ã®ãƒ‡ãƒ¼ã‚¿èª­ã¿è¾¼ã¿ä¸­ã«äºˆæœŸã›ã¬ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {e}")
        st.exception(e)
        st.stop()
    return []

# Excelãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆã«ãƒ•ã‚©ãƒ¼ãƒ ãƒ‡ãƒ¼ã‚¿ã‚’æ›¸ãè¾¼ã‚€é–¢æ•°
def generate_excel_from_form_data(form_data):
    output_excel = BytesIO()
    try:
        # æ—¢å­˜ã®Excelãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆãƒ•ã‚¡ã‚¤ãƒ«ï¼ˆæˆæ¥­ã‚«ãƒ¼ãƒ‰.xlsmï¼‰ã‚’èª­ã¿è¾¼ã‚€
        # openpyxlã¯xlsmãƒ•ã‚¡ã‚¤ãƒ«ã‚’æ‰±ãˆã¾ã™ãŒã€VBAãƒã‚¯ãƒ­ã¯ä¿æŒã™ã‚‹ã‚‚ã®ã®å®Ÿè¡Œã¯ã§ãã¾ã›ã‚“ã€‚
        # VBAãƒã‚¯ãƒ­ã‚’å®Œå…¨ã«æ©Ÿèƒ½ã•ã›ãŸã„å ´åˆã¯ã€Pythonã‹ã‚‰å¤–éƒ¨ãƒ„ãƒ¼ãƒ«ã‚’å‘¼ã³å‡ºã™ãªã©è¤‡é›‘ãªå¯¾å¿œãŒå¿…è¦ã§ã™ã€‚
        # ä»Šå›ã¯ãƒ‡ãƒ¼ã‚¿ã‚’æ›¸ãè¾¼ã‚€ã“ã¨ã®ã¿ã«ç„¦ç‚¹ã‚’å½“ã¦ã¾ã™ã€‚
        with open("æˆæ¥­ã‚«ãƒ¼ãƒ‰.xlsm", "rb") as f:
            workbook_data = BytesIO(f.read())
        
        wb = load_workbook(workbook_data, read_only=False, keep_vba=True)
        ws = wb.active # ã‚¢ã‚¯ãƒ†ã‚£ãƒ–ãªã‚·ãƒ¼ãƒˆ

        # Googleãƒ•ã‚©ãƒ¼ãƒ ã®é …ç›®åï¼ˆã‚¹ãƒ—ãƒ¬ãƒƒãƒ‰ã‚·ãƒ¼ãƒˆã®ãƒ˜ãƒƒãƒ€ãƒ¼åï¼‰ã¨Excelã®ã‚»ãƒ«ä½ç½®ã‚’å¯¾å¿œã•ã›ã‚‹
        # ã“ã®ãƒãƒƒãƒ”ãƒ³ã‚°ã¯ã€ã‚ãªãŸã®ã€Œæˆæ¥­ã‚«ãƒ¼ãƒ‰.xlsmã€ã®ãƒ¬ã‚¤ã‚¢ã‚¦ãƒˆã«åˆã‚ã›ã¦èª¿æ•´ã—ã¦ãã ã•ã„ã€‚
        # ä¾‹: Googleãƒ•ã‚©ãƒ¼ãƒ ã®é …ç›®å -> Excelã®ã‚»ãƒ«
        cell_mappings = {
            'ã‚¿ã‚¤ãƒ ã‚¹ã‚¿ãƒ³ãƒ—': 'A2', # ã‚¹ãƒ—ãƒ¬ãƒƒãƒ‰ã‚·ãƒ¼ãƒˆã®Aåˆ—
            'å˜å…ƒå': 'B2',
            'ã‚­ãƒ£ãƒƒãƒã‚³ãƒ”ãƒ¼': 'C2',
            'ã­ã‚‰ã„': 'D2',
            'å¯¾è±¡å­¦å¹´': 'E2',
            'éšœå®³ç¨®åˆ¥': 'F2',
            'æ™‚é–“': 'G2',
            'æº–å‚™ç‰©': 'H2',
            'å°å…¥ã®æµã‚Œ': 'I2',
            'æ´»å‹•ã®æµã‚Œ': 'J2',
            'æŒ¯ã‚Šè¿”ã‚Šã®æµã‚Œ': 'K2',
            'æŒ‡å°ã®ãƒã‚¤ãƒ³ãƒˆ': 'L2',
            'ãƒãƒƒã‚·ãƒ¥ã‚¿ã‚°': 'M2',
            'ãƒ¡ã‚¤ãƒ³ç”»åƒURL': 'N2',
            'æ•™æå†™çœŸURL': 'O2',
            'å‹•ç”»ãƒªãƒ³ã‚¯': 'P2',
            'æŒ‡å°æ¡ˆWordãƒ•ã‚¡ã‚¤ãƒ«URL': 'Q2',
            'æŒ‡å°æ¡ˆPDFãƒ•ã‚¡ã‚¤ãƒ«URL': 'R2',
            'æˆæ¥­è³‡æ–™PowerPointãƒ•ã‚¡ã‚¤ãƒ«URL': 'S2',
            'è©•ä¾¡ã‚·ãƒ¼ãƒˆExcelãƒ•ã‚¡ã‚¤ãƒ«URL': 'T2',
            'ICTæ´»ç”¨æœ‰ç„¡': 'U2',
            'æ•™ç§‘': 'V2',
            'å­¦ç¿’é›†å›£ã®å˜ä½': 'W2',
            'å˜å…ƒå†…ã®æˆæ¥­ã‚¿ã‚¤ãƒˆãƒ«': 'X2',
            'å˜å…ƒå†…ã§ã®ä¸¦ã³é †': 'Y2', # æ–°è¦è¿½åŠ ã€å¿…è¦ã«å¿œã˜ã¦
        }

        for form_field, excel_cell in cell_mappings.items():
            value = form_data.get(form_field, '')
            if isinstance(value, list):
                # ãƒªã‚¹ãƒˆã®å ´åˆã¯ã‚»ãƒŸã‚³ãƒ­ãƒ³ã¾ãŸã¯ã‚«ãƒ³ãƒã§çµåˆï¼ˆç”¨é€”ã«ã‚ˆã‚‹ï¼‰
                if form_field == 'ãƒãƒƒã‚·ãƒ¥ã‚¿ã‚°':
                    ws[excel_cell] = ','.join(value)
                else:
                    ws[excel_cell] = ';'.join(value)
            else:
                ws[excel_cell] = str(value)
        
        wb.save(output_excel)
        output_excel.seek(0)
        return output_excel.getvalue()

    except FileNotFoundError:
        st.error("âš ï¸ 'æˆæ¥­ã‚«ãƒ¼ãƒ‰.xlsm' ãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆãƒ•ã‚¡ã‚¤ãƒ«ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã€‚ã‚¢ãƒ—ãƒªã‚±ãƒ¼ã‚·ãƒ§ãƒ³ã¨åŒã˜éšå±¤ã«é…ç½®ã—ã¦ãã ã•ã„ã€‚")
        return None
    except Exception as e:
        st.error(f"Excelãƒ•ã‚¡ã‚¤ãƒ«ã®ç”Ÿæˆä¸­ã«ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {e}")
        st.exception(e)
        return None

# --- Streamlit UI ---

st.info("""
ã“ã®ã‚¢ãƒ—ãƒªã§ã¯ã€Googleãƒ•ã‚©ãƒ¼ãƒ ã§å…¥åŠ›ã•ã‚ŒãŸå›ç­”ã‚’ã‚‚ã¨ã«ã€å€‹åˆ¥ã®æˆæ¥­ã‚«ãƒ¼ãƒ‰Excelãƒ•ã‚¡ã‚¤ãƒ«ã‚’ç”Ÿæˆãƒ»ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ã§ãã¾ã™ã€‚
åˆ©ç”¨ã™ã‚‹ãŸã‚ã«ã¯ã€`secrets.toml` ãƒ•ã‚¡ã‚¤ãƒ«ã¾ãŸã¯Streamlit Cloudã®ã‚·ãƒ¼ã‚¯ãƒ¬ãƒƒãƒˆè¨­å®šãŒå¿…è¦ã§ã™ã€‚
""")

# Google Sheets API ã®è¨­å®šã‚’ secrets ã‹ã‚‰å–å¾—
GOOGLE_SHEET_SPREADSHEET_NAME = st.secrets.get("google_sheet_spreadsheet_name", "ã‚ãªãŸã®ã‚¹ãƒ—ãƒ¬ãƒƒãƒ‰ã‚·ãƒ¼ãƒˆå")
GOOGLE_SHEET_WORKSHEET_NAME = st.secrets.get("google_sheet_worksheet_name", "ãƒ•ã‚©ãƒ¼ãƒ ã®å›ç­” 1")

# Googleãƒ•ã‚©ãƒ¼ãƒ ã®URL (ãƒ¦ãƒ¼ã‚¶ãƒ¼ã«è¡¨ç¤ºã™ã‚‹ãŸã‚ã®ãƒ€ãƒŸãƒ¼ãƒªãƒ³ã‚¯)
# å®Ÿéš›ã«ä½¿ç”¨ã™ã‚‹Googleãƒ•ã‚©ãƒ¼ãƒ ã®URLã«ç½®ãæ›ãˆã¦ãã ã•ã„ã€‚
google_form_input_link = st.secrets.get("google_form_url", "https://forms.gle/YOUR_ACTUAL_GOOGLE_FORM_LINK")

st.markdown(
    f"""
    <p style="font-size:1.1em;">å…¥åŠ›ç”¨ã®Googleãƒ•ã‚©ãƒ¼ãƒ ã¯ã“ã¡ã‚‰: <a href="{google_form_input_link}" target="_blank">ğŸ“ Googleãƒ•ã‚©ãƒ¼ãƒ ã‚’é–‹ã</a></p>
    """, unsafe_allow_html=True
)

if google_form_input_link == "https://forms.gle/YOUR_ACTUAL_GOOGLE_FORM_LINK":
    st.warning("âš ï¸ Googleãƒ•ã‚©ãƒ¼ãƒ ã®URLã‚’ã€Streamlitã®secrets (`google_form_url` ã‚­ãƒ¼) ã¾ãŸã¯ã‚³ãƒ¼ãƒ‰å†…ã§å®Ÿéš›ã®URLã«æ›´æ–°ã—ã¦ãã ã•ã„ã€‚")

# æœ€æ–°ã®ãƒ•ã‚©ãƒ¼ãƒ å›ç­”ã‚’èª­ã¿è¾¼ã‚€ãƒœã‚¿ãƒ³
if st.button("ğŸ”„ æœ€æ–°ã®ãƒ•ã‚©ãƒ¼ãƒ å›ç­”ã‚’èª­ã¿è¾¼ã‚€"):
    load_data_from_google_sheet.clear() # ã‚­ãƒ£ãƒƒã‚·ãƒ¥ã‚’ã‚¯ãƒªã‚¢ã—ã¦å†èª­ã¿è¾¼ã¿ã‚’å¼·åˆ¶
    sheet_lesson_data_records = load_data_from_google_sheet(
        spreadsheet_name=GOOGLE_SHEET_SPREADSHEET_NAME,
        worksheet_name=GOOGLE_SHEET_WORKSHEET_NAME
    )
    if sheet_lesson_data_records:
        st.session_state.google_form_records = sheet_lesson_data_records
        st.success(f"{len(sheet_lesson_data_records)}ä»¶ã®Googleãƒ•ã‚©ãƒ¼ãƒ å›ç­”ã‚’èª­ã¿è¾¼ã¿ã¾ã—ãŸï¼")
    else:
        st.info("Googleãƒ•ã‚©ãƒ¼ãƒ ã‹ã‚‰ã®å›ç­”ãŒè¦‹ã¤ã‹ã‚Šã¾ã›ã‚“ã§ã—ãŸã€‚")

# Googleãƒ•ã‚©ãƒ¼ãƒ ã‹ã‚‰å–å¾—ã—ãŸãƒ‡ãƒ¼ã‚¿ãŒã‚ã‚‹å ´åˆã€é¸æŠãƒ»ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰UIã‚’è¡¨ç¤º
if 'google_form_records' in st.session_state and st.session_state.google_form_records:
    
    st.markdown("---")
    st.subheader("â¬‡ï¸ Excelæˆæ¥­ã‚«ãƒ¼ãƒ‰ç”Ÿæˆ")

    # ãƒ‰ãƒ­ãƒƒãƒ—ãƒ€ã‚¦ãƒ³ãƒªã‚¹ãƒˆã«è¡¨ç¤ºã™ã‚‹ã‚ªãƒ—ã‚·ãƒ§ãƒ³ã‚’ä½œæˆ
    # ä¾‹: "{å˜å…ƒå} - {ã‚¿ã‚¤ãƒ ã‚¹ã‚¿ãƒ³ãƒ—}"
    selection_options = [
        f"[{entry.get('ã‚¿ã‚¤ãƒ ã‚¹ã‚¿ãƒ³ãƒ—', 'æ—¥æ™‚ä¸æ˜')}] {entry.get('å˜å…ƒå', 'å˜å…ƒåãªã—')} - {entry.get('å˜å…ƒå†…ã®æˆæ¥­ã‚¿ã‚¤ãƒˆãƒ«', 'æˆæ¥­ã‚¿ã‚¤ãƒˆãƒ«ãªã—')}"
        for entry in st.session_state.google_form_records
    ]
    
    selected_index = st.selectbox(
        "ExcelåŒ–ã™ã‚‹ãƒ•ã‚©ãƒ¼ãƒ å›ç­”ã‚’é¸æŠã—ã¦ãã ã•ã„",
        options=range(len(selection_options)),
        format_func=lambda x: selection_options[x],
        key="selected_form_entry_for_excel"
    )

    if selected_index is not None:
        selected_form_entry = st.session_state.google_form_records[selected_index]
        
        # Excelç”Ÿæˆãƒœã‚¿ãƒ³
        if st.button(f"ã€Œ{selected_form_entry.get('å˜å…ƒå', 'é¸æŠã•ã‚ŒãŸå›ç­”')}ã€ã®Excelã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰", key="download_generated_excel"):
            excel_data = generate_excel_from_form_data(selected_form_entry)
            if excel_data:
                # ãƒ•ã‚¡ã‚¤ãƒ«åã«å˜å…ƒåã¨ã‚¿ã‚¤ãƒ ã‚¹ã‚¿ãƒ³ãƒ—ã‚’å«ã‚ã‚‹
                unit_name_for_filename = selected_form_entry.get('å˜å…ƒå', 'æˆæ¥­ã‚«ãƒ¼ãƒ‰').replace(' ', '_').replace('/', '_')
                timestamp_for_filename = selected_form_entry.get('ã‚¿ã‚¤ãƒ ã‚¹ã‚¿ãƒ³ãƒ—', '').split(' ')[0].replace('-', '') # æ—¥ä»˜ã®ã¿ã‚’ä½¿ç”¨
                download_filename = f"{unit_name_for_filename}_æˆæ¥­ã‚«ãƒ¼ãƒ‰_{timestamp_for_filename}.xlsm"
                
                st.download_button(
                    label="âœ… Excelãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰",
                    data=excel_data,
                    file_name=download_filename,
                    mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                    help="Googleãƒ•ã‚©ãƒ¼ãƒ ã®å›ç­”ã‹ã‚‰ç”Ÿæˆã•ã‚ŒãŸExcelãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ã—ã¾ã™ã€‚"
                )
                st.success("æŒ‡å®šã•ã‚ŒãŸãƒ•ã‚©ãƒ¼ãƒ å›ç­”ã‹ã‚‰Excelãƒ•ã‚¡ã‚¤ãƒ«ã‚’ç”Ÿæˆã—ã¾ã—ãŸï¼")
            else:
                st.error("Excelãƒ•ã‚¡ã‚¤ãƒ«ã®ç”Ÿæˆã«å¤±æ•—ã—ã¾ã—ãŸã€‚ä¸Šè¨˜ã®ã‚¨ãƒ©ãƒ¼ãƒ¡ãƒƒã‚»ãƒ¼ã‚¸ã‚’ç¢ºèªã—ã¦ãã ã•ã„ã€‚")
    else:
        st.info("é¸æŠå¯èƒ½ãªãƒ•ã‚©ãƒ¼ãƒ å›ç­”ãŒã‚ã‚Šã¾ã›ã‚“ã€‚")
else:
    st.info("ã¾ãšã€Œæœ€æ–°ã®ãƒ•ã‚©ãƒ¼ãƒ å›ç­”ã‚’èª­ã¿è¾¼ã‚€ã€ãƒœã‚¿ãƒ³ã‚’æŠ¼ã—ã¦ã€Googleãƒ•ã‚©ãƒ¼ãƒ ã®ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—ã—ã¦ãã ã•ã„ã€‚")

st.markdown("---")
st.markdown("### â„¹ï¸ è¨­å®šã‚¬ã‚¤ãƒ‰")
st.markdown("""
ã“ã®ã‚¢ãƒ—ãƒªã‚’å‹•ä½œã•ã›ã‚‹ã«ã¯ã€Streamlitã® `secrets.toml` ã¾ãŸã¯Streamlit Cloudã®ã‚·ãƒ¼ã‚¯ãƒ¬ãƒƒãƒˆè¨­å®šã«ä»¥ä¸‹ã®æƒ…å ±ãŒå¿…è¦ã§ã™ã€‚

1.  **`GOOGLE_SHEETS_CREDENTIALS`**: Googleã‚µãƒ¼ãƒ“ã‚¹ã‚¢ã‚«ã‚¦ãƒ³ãƒˆã®JSONã‚­ãƒ¼ã‚’ãã®ã¾ã¾æ–‡å­—åˆ—ã¨ã—ã¦è²¼ã‚Šä»˜ã‘ã¾ã™ã€‚
    ```toml
    GOOGLE_SHEETS_CREDENTIALS = '''
    {
      "type": "service_account",
      "project_id": "your-project-id",
      "private_key_id": "...",
      "private_key": "-----BEGIN PRIVATE KEY-----\\n...\\n-----END PRIVATE KEY-----\\n",
      "client_email": "...",
      "client_id": "...",
      "auth_uri": "...",
      "token_uri": "...",
      "auth_provider_x509_cert_url": "...",
      "client_x509_cert_url": "...",
      "universe_domain": "..."
    }
    '''
    ```
    **æ³¨æ„**: `private_key` ã®æ”¹è¡Œæ–‡å­— `\\n` ã‚’å¿˜ã‚Œãªã„ã§ãã ã•ã„ã€‚
2.  **`google_sheet_spreadsheet_name`**: Googleã‚¹ãƒ—ãƒ¬ãƒƒãƒ‰ã‚·ãƒ¼ãƒˆã®æ­£ç¢ºãªåå‰ (ä¾‹: "æˆæ¥­ã‚«ãƒ¼ãƒ‰ ï¼ˆå›ç­”ï¼‰")
    ```toml
    google_sheet_spreadsheet_name = "æˆæ¥­ã‚«ãƒ¼ãƒ‰ ï¼ˆå›ç­”ï¼‰"
    ```
3.  **`google_sheet_worksheet_name`**: ãƒ•ã‚©ãƒ¼ãƒ ã®å›ç­”ãŒè¨˜éŒ²ã•ã‚Œã¦ã„ã‚‹ãƒ¯ãƒ¼ã‚¯ã‚·ãƒ¼ãƒˆã®åå‰ (ä¾‹: "ãƒ•ã‚©ãƒ¼ãƒ ã®å›ç­” 1")
    ```toml
    google_sheet_worksheet_name = "ãƒ•ã‚©ãƒ¼ãƒ ã®å›ç­” 1"
    ```
4.  **`google_form_url` (ã‚ªãƒ—ã‚·ãƒ§ãƒ³)**: ãƒ¦ãƒ¼ã‚¶ãƒ¼ã«è¡¨ç¤ºã™ã‚‹Googleãƒ•ã‚©ãƒ¼ãƒ ã®URLã€‚
    ```toml
    google_form_url = "https://forms.gle/YOUR_ACTUAL_GOOGLE_FORM_LINK"
    ```

ã¾ãŸã€`æˆæ¥­ã‚«ãƒ¼ãƒ‰.xlsm` ã¨ã„ã†åå‰ã®Excelãƒ†ãƒ³ãƒ—ãƒ¬ãƒ¼ãƒˆãƒ•ã‚¡ã‚¤ãƒ«ãŒã€ã“ã®Pythonã‚¹ã‚¯ãƒªãƒ—ãƒˆã¨åŒã˜ãƒ‡ã‚£ãƒ¬ã‚¯ãƒˆãƒªã«å­˜åœ¨ã™ã‚‹å¿…è¦ãŒã‚ã‚Šã¾ã™ã€‚
""")